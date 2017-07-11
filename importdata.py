import openpyxl
from openpyxl import load_workbook
from classproject import settings
from django.core.management.base import BaseCommand, CommandError
import MySQLdb

class Command(BaseCommand):

    def handle(self, *args, **options):
        db = MySQLdb.connect("localhost", "testuser", "test123", "TESTDB")
        cursor = db.cursor()
        source = "C:\Users\karanam\classproject\students.xlsx"
        sr = load_workbook(source)
        sheets = sr.get_sheet_names()
        ss = sr.get_sheet_by_name("Current")
        for i in xrange(1, ss.max_row + 1):
            name = ss.cell(row=i, column=1).value
            college =  ss.cell(row=i, column=2).value
            email =  ss.cell(row=i, column=3).value
            db_folder =  ss.cell(row=i, column=4).value
            sql = "INSERT INTO STUDENT(name, \
                college, email, db_folder) \
                VALUES ('%s', '%s', '%s', '%s' )" % \
                  (name, college, email,db_folder)
            cursor.execute(sql)
        db.close()