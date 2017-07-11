
import os
from django.core.wsgi import get_wsgi_application
os.environ['DJANGO_SETTINGS_MODULE'] = 'classproject.settings'
application = get_wsgi_application()
from classproject import settings
from onlineapp import models
import click
from openpyxl import load_workbook
from openpyxl import Workbook
from bs4 import BeautifulSoup
import warnings
warnings.filterwarnings('ignore')
import  MySQLdb
from MySQLdb import Error

@click.group()
def cli():
    pass

@cli.command()
def createdb() :
    try:
        db = MySQLdb.connect("localhost", settings.DATABASES['default']['USER'], settings.DATABASES['default']['PASSWORD'], port = 3308)
        cursor = db.cursor()
        sql = 'CREATE SCHEMA IF NOT EXISTS ' + settings.DATABASES['default']['NAME']
        cursor.execute(sql)
        db.commit()
    except Error as e :
             print e

@cli.command()
def dropdb():
    try :
        db = MySQLdb.connect("localhost", settings.DATABASES['default']['USER'], settings.DATABASES['default']['PASSWORD'],port = 3308)
        cursor = db.cursor()
        sql = 'DROP DATABASE ' + settings.DATABASES['default']['NAME']
        cursor.execute(sql)
        db.commit()
    except Error as e:
        print e


@cli.command()
@click.argument('excelfile',nargs=1)
@click.argument('htmlfile',nargs=1)
def populatedb(excelfile,htmlfile) :
    if excelfile:
        db = MySQLdb.connect("localhost","root","raghu", "classdb", 3308)
        wb = load_workbook(excelfile)
        ws = wb['Colleges']
        data = map(lambda x: {'name': x[0].value,
                              'acronym': x[1].value,
                              'location': x[2].value,
                              'contact': x[3].value.lower()},
                   ws[2: ws.max_row])
        print len(data)
        for i in range(len(data)):
            c= models.College.objects.create(name = data[i]['name'],location =data[i]['location'] , acronym = data[i]['acronym'],contact =data[i]['contact'])
            c.save()
    if excelfile:
        db = MySQLdb.connect("localhost","root", "raghu","classdb", 3308)
        wb = load_workbook(excelfile)
        ws = wb['Current']

        data = map(lambda x: {'name': x[0].value,
                              'college': x[1].value,
                              'contact': x[2].value,
                              'dbnames': x[3].value.lower()},
                   ws[2: ws.max_row])
        for i in range(len(data)):
            j = models.College.objects.get(acronym=data[i]['college'])
            c = models.Student.objects.create(name=data[i]['name'], college=j,
                                              email=data[i]['contact'], db_folder=data[i]['dbnames'])
            c.save()

        ws = wb['Deletions']
        data = map(lambda x: {'name': x[0].value,
                              'college': x[1].value,
                              'contact': x[2].value,
                              'dbnames': x[3].value.lower()},
                   ws[2: ws.max_row])
        for i in range(len(data)):
            try:
                j = models.College.objects.get(acronym=data[i]['college'])
                c = models.Student.objects.create(name=data[i]['name'], college=j,
                                                  email=data[i]['contact'], db_folder=data[i]['dbnames'],dropped_out=1)
                c.save()
            except:
                continue
    soup = BeautifulSoup(open(htmlfile),'html.parser')
    table = soup.find('table')
    row = table.findAll("tr")
    wb = Workbook()
    ws = wb.active
    for i in row[1:]:
        a=[]
        for j in i.findAll('td')[1:]:
                a.append(j.string)
        try:
            obj = models.Student.objects.get(db_folder= a[0].split('_')[2])
            c= models.MockTest1(problem1=a[1],problem2=a[2],problem3=a[3],problem4=a[4],total= a[5],student=obj)
            c.save()
        except:
            continue

if __name__ == "__main__" :
    cli()
